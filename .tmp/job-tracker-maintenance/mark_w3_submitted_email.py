from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path(
    "/Users/yuminseog/portfolio/outputs/job-application-tracker/"
    "ai_creative_marketing_application_tracker_2026-06-15.xlsx"
)

POSTING_URL = "https://www.jobkorea.co.kr/Recruit/GI_Read/49244888"
MESSAGE_ID = "19edf86aa3286ed8"
SUBMITTED_AT = datetime(2026, 6, 19)

EMAIL_BODY = """안녕하세요, 더블유쓰리컴퍼니 AI 콘텐츠 마케터 인턴 포지션에 지원하는 유민석입니다.

저는 AI로 숏폼 광고, AI 아이돌/IP 콘텐츠, 제품 공개/마케팅 콘텐츠를 기획하고 실제 산출물로 제작해왔습니다. Moji는 글로벌 친구·언어교환 중심의 가벼운 매칭 서비스, SeriUs는 검증과 신뢰 기반의 serious international dating 서비스로 이해했고, 두 서비스의 톤을 나누어 AI 숏폼·이미지·오디오 광고 소재를 빠르게 실험하는 역할에 강하게 맞는다고 판단해 지원드립니다.

영문 CV와 포트폴리오 링크를 함께 전달드립니다.
- Portfolio: https://portfolio-yuminseoks-projects.vercel.app
- Loom/Pulso: https://loom-signal-deck.vercel.app/
- AHEYA: https://aheyabaraya.xyz/

감사합니다.
유민석 드림
+82 10-4229-9239
aheayabaraya@gmail.com"""


def header_map(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(header_row, col).value
    }


def append_note(existing: str | None, note: str) -> str:
    if not existing:
        return note
    if note in existing:
        return existing
    return f"{existing} / {note}"


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def main() -> None:
    wb = load_workbook(WORKBOOK)

    manage = wb["지원관리"]
    manage_headers = header_map(manage, 8)
    for row in range(9, manage.max_row + 1):
        company = manage.cell(row, manage_headers["회사"]).value
        role = manage.cell(row, manage_headers["공고명"]).value
        if company == "더블유쓰리컴퍼니" and role == "AI 콘텐츠 마케터 (인턴)":
            manage.cell(row, manage_headers["우선"]).value = "완료"
            manage.cell(row, manage_headers["상태"]).value = "지원완료"
            manage.cell(row, manage_headers["요구자료"]).value = (
                "제출 완료: 영문 CV PDF 첨부, 포트폴리오 링크 3개 메일 본문 삽입"
            )
            manage.cell(row, manage_headers["다음 액션"]).value = "결과 대기"
            manage.cell(row, manage_headers["분류 판단"]).value = (
                "Gmail 이메일 지원 완료. Moji/SeriUs 서비스 톤 구분, "
                "AI content creative experimentation, short-form/social ad workflow 중심으로 제출."
            )
            manage.cell(row, manage_headers["업데이트"]).value = SUBMITTED_AT
            manage.cell(row, manage_headers["비고"]).value = append_note(
                manage.cell(row, manage_headers["비고"]).value,
                f"2026-06-19 Gmail 발송 완료. message id {MESSAGE_ID}. "
                "첨부: Minseok_Yu_AI_Content_Marketer_CV.pdf",
            )
            manage.cell(row, manage_headers["체크"]).value = "완료"
            break
    else:
        raise RuntimeError("W3Company row not found in 지원관리")

    drafts = wb["회사별자소서"]
    draft_headers = header_map(drafts, 4)

    for row in range(5, drafts.max_row + 1):
        if drafts.cell(row, draft_headers["회사"]).value == "더블유쓰리컴퍼니":
            drafts.cell(row, draft_headers["상태"]).value = "제출반영"
            drafts.cell(row, draft_headers["수정일"]).value = SUBMITTED_AT
            drafts.cell(row, draft_headers["비고"]).value = append_note(
                drafts.cell(row, draft_headers["비고"]).value,
                "2026-06-19 실제 이메일 지원에 반영",
            )

    email_row = None
    for row in range(5, drafts.max_row + 1):
        if (
            drafts.cell(row, draft_headers["회사"]).value == "더블유쓰리컴퍼니"
            and drafts.cell(row, draft_headers["섹션"]).value == "이메일 본문"
        ):
            email_row = row
            break

    if email_row is None:
        email_row = drafts.max_row + 1
        copy_row_style(drafts, 11, email_row)

    values = {
        "상태": "제출반영",
        "회사": "더블유쓰리컴퍼니",
        "공고명": "AI 콘텐츠 마케터 (인턴)",
        "섹션": "이메일 본문",
        "회사맞춤 포지셔닝": "Moji/SeriUs 서비스 이해 + AI Content Marketer 지원 메일",
        "완성문": EMAIL_BODY,
        "글자수": len(EMAIL_BODY),
        "사용 포트폴리오": "MUSINSA, Loom/Pulso, AHEYA",
        "제출처": "Gmail 이메일",
        "공고 URL": POSTING_URL,
        "작성일": SUBMITTED_AT,
        "수정일": SUBMITTED_AT,
        "비고": (
            f"2026-06-19 실제 발송. Gmail message id {MESSAGE_ID}. "
            "첨부: Minseok_Yu_AI_Content_Marketer_CV.pdf"
        ),
    }
    for key, value in values.items():
        drafts.cell(email_row, draft_headers[key]).value = value

    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")


if __name__ == "__main__":
    main()
